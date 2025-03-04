import sys
import os
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.chart import BarChart, Reference

def add_summary(master_filename):
    """
    Ajoute une feuille 'Résumé' avec des totaux et un graphique.
    - Deux colonnes : 'Catégorie' et 'Nombre'
    - Évite les erreurs de tableaux Excel corrompus
    """

    if not os.path.exists(master_filename):
        print(f"[ERREUR] Le fichier '{master_filename}' n'existe pas.")
        return

    print("\n[INFO] Création d'une feuille 'Résumé' avec des totaux et un graphique dans le fichier Master...")

    # Charger le fichier Excel
    wb = load_workbook(master_filename)

    # Suppression propre de la feuille "Résumé" si elle existe déjà
    if "Résumé" in wb.sheetnames:
        del wb["Résumé"]
    ws_summary = wb.create_sheet("Résumé", 0)

    # Définition des en-têtes
    headers = ["Catégorie", "Nombre"]
    ws_summary.append(headers)

    # Appliquer un style aux en-têtes
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    for col_num, header in enumerate(headers, start=1):
        col_letter = get_column_letter(col_num)
        cell = ws_summary[f"{col_letter}1"]
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
        cell.fill = header_fill

    # Exemple de données (à remplacer par les données réelles)
    data = [
        ["Bug Fixes", 10],
        ["Enhancements", 5],
        ["Documentation", 3]
    ]

    for row in data:
        ws_summary.append(row)

    # Ajustement des colonnes
    ws_summary.column_dimensions["A"].width = 25
    ws_summary.column_dimensions["B"].width = 15

    # Ajout d'un graphique à barres
    chart = BarChart()
    chart.title = "Résumé des Catégories"
    chart.x_axis.title = "Catégorie"
    chart.y_axis.title = "Nombre"

    data_ref = Reference(ws_summary, min_col=2, min_row=1, max_row=len(data) + 1)
    categories_ref = Reference(ws_summary, min_col=1, min_row=2, max_row=len(data) + 1)
    chart.add_data(data_ref, titles_from_data=True)
    chart.set_categories(categories_ref)
    chart.shape = 4
    ws_summary.add_chart(chart, "D4")

    # Sauvegarde propre sans tableau
    wb.save(master_filename)
    print(f"[SUCCÈS] Feuille 'Résumé' ajoutée à '{master_filename}' avec succès, avec des totaux et un graphique.")

if __name__ == "__main__":
    # Vérifier si un fichier Master a été passé en argument
    if len(sys.argv) < 2:
        print("[ERREUR] Aucun fichier Master spécifié.")
        sys.exit(1)

    master_file = sys.argv[1]
    add_summary(master_file)
