import sys
import argparse
from bs4 import BeautifulSoup
import pandas as pd
from openpyxl import load_workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter

# Définition des sections à extraire
SECTIONS_MAPPING = {
    "correctifs": [
        "Anomalies corrigées du produit",
        "Anomalies corrigées de la configuration par défaut",
        "Anomalies corrigées de Vantage",
        "Anomalies corrigées de Deliver"
    ],
    "améliorations": [
        "Améliorations du produit",
        "Améliorations de la configuration par défaut",
        "Améliorations de Deliver",
        "Améliorations des intégrations"
    ]
}

def extract_data(html_file, output_file, extract_type):
    """
    Extrait les correctifs ou les améliorations en fonction du type spécifié.
    """
    print(f"[INFO] Analyse du fichier : {html_file}")
    print(f"[INFO] Fichier Excel de sortie : {output_file}")

    # Charger le fichier HTML
    with open(html_file, "r", encoding="utf-8") as file:
        soup = BeautifulSoup(file, "html.parser")

    # Définir les sections à rechercher selon le type
    if extract_type not in SECTIONS_MAPPING:
        print(f"[ERREUR] Type d'extraction invalide : {extract_type}. Utilisez 'correctifs' ou 'améliorations'.")
        sys.exit(1)

    sections_to_find = SECTIONS_MAPPING[extract_type]

    # Trouver dynamiquement les sections par leur titre
    sections_found = {}
    for h in soup.find_all(["h1", "h2"]):  # Recherche dans h1 et h2
        section_title = h.get_text(strip=True)
        for target_section in sections_to_find:
            if target_section.lower() in section_title.lower():  # Comparaison insensible à la casse
                sections_found[target_section] = h
                break  # Passer à la section suivante dès qu'on a un match

    # Vérifier si toutes les sections attendues ont été trouvées
    missing_sections = [s for s in sections_to_find if s not in sections_found]
    if missing_sections:
        print(f"[AVERTISSEMENT] Certaines sections sont introuvables : {missing_sections}")

    # Extraction des données
    data = []
    for section, header_tag in sections_found.items():
        subsections = header_tag.find_all_next(["h2", "h3", "p"], limit=None)

        current_subsection = None
        issue_id = None
        for tag in subsections:
            if tag.name == "h2":
                current_subsection = tag.text.strip()
            elif tag.name == "h3":
                issue_id = tag.text.strip()
            elif tag.name == "p":
                description = tag.text.strip()
                if current_subsection and issue_id and description:
                    data.append([section, current_subsection, issue_id, description])

    # Vérifier si des données ont été extraites
    if not data:
        print(f"[AVERTISSEMENT] Aucune donnée trouvée dans '{html_file}'. Vérifiez que les sections existent dans la page.")
        sys.exit(1)

    # Création d'un DataFrame
    df = pd.DataFrame(data, columns=["Section", "Sous-section", "ID", "Description"])

    # Suppression des doublons sur "ID" uniquement si la "Description" est vide
    df = df[~((df.duplicated(subset=["ID"], keep=False)) & (df["Description"] == ""))]

    # Ajouter une colonne "Test Status" vide
    df["Test Status"] = ""

    # Sauvegarde des données en Excel
    df.to_excel(output_file, index=False)

    # Appliquer un formatage avec OpenPyXL
    format_excel(output_file)

def format_excel(file_path):
    """
    Applique un formatage Excel :
    - Convertit le fichier en tableau structuré
    - Ajuste les colonnes
    - Ajoute des styles pour la lisibilité
    """
    wb = load_workbook(file_path)
    ws = wb.active

    num_rows = ws.max_row
    num_cols = ws.max_column

    if num_rows < 2 or num_cols < 1:  # Vérifier que le fichier contient des données
        print(f"[AVERTISSEMENT] Le fichier '{file_path}' est vide après extraction, aucun tableau ajouté.")
    else:
        col_letter = get_column_letter(num_cols)
        range_ref = f"A1:{col_letter}{num_rows}"

        # Création du tableau formaté
        table = Table(displayName="PatchNotesTable", ref=range_ref)
        style = TableStyleInfo(
            name="TableStyleMedium9",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        table.tableStyleInfo = style
        ws.add_table(table)
        print(f"[INFO] Tableau formaté ajouté dans '{file_path}'.")

    # Sauvegarde finale
    wb.save(file_path)
    print(f"[SUCCÈS] Fichier '{file_path}' généré avec succès avec un tableau formaté.")

if __name__ == "__main__":
    # Parsing des arguments
    parser = argparse.ArgumentParser(description="Extraction des patch notes de Sciforma")
    parser.add_argument("html_file", help="Fichier HTML contenant les patch notes")
    parser.add_argument("--type", choices=["correctifs", "améliorations"], required=True, help="Type d'extraction")
    parser.add_argument("output_file", help="Nom du fichier Excel de sortie")

    args = parser.parse_args()

    # Exécuter l'extraction avec les arguments
    extract_data(args.html_file, args.output_file, args.type)
