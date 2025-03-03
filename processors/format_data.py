import sys
import os
from openpyxl import load_workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Font

def format_excel(file_path):
    """
    Applique un formatage automatique au fichier Excel :
    - Convertit les données en tableau structuré
    - Ajuste la largeur des colonnes
    - Applique du Word Wrap sur toutes les cellules
    - Centre les en-têtes en gras
    """

    if not os.path.exists(file_path):
        print(f"[ERREUR] Le fichier '{file_path}' n'existe pas.")
        return

    print(f"[INFO] Mise en forme du fichier : {file_path}")

    # Charger le fichier Excel
    wb = load_workbook(file_path)

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        num_rows = ws.max_row
        num_cols = ws.max_column

        if num_rows < 2 or num_cols < 1:
            print(f"[AVERTISSEMENT] Feuille '{sheet_name}' ignorée car elle est vide.")
            continue

        col_letter = get_column_letter(num_cols)
        range_ref = f"A1:{col_letter}{num_rows}"

        # Convertir en tableau Excel
        table = Table(displayName=f"{sheet_name}Table", ref=range_ref)
        style = TableStyleInfo(
            name="TableStyleMedium9",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        table.tableStyleInfo = style
        ws.add_table(table)

        # Ajustement automatique des colonnes + Word Wrap
        for col in range(1, num_cols + 1):
            col_letter = get_column_letter(col)

            if col == 4:  # Colonne "Description"
                ws.column_dimensions[col_letter].width = 50  # Largeur plus grande pour descriptions
            else:
                max_length = max((len(str(ws[f"{col_letter}{row}"].value)) for row in range(1, num_rows + 1)), default=10)
                ws.column_dimensions[col_letter].width = min(max_length + 2, 30)  # Largeur max de 30

            # Appliquer Word Wrap et alignement à toutes les cellules
            for row in range(1, num_rows + 1):
                cell = ws[f"{col_letter}{row}"]
                cell.alignment = Alignment(wrap_text=True, vertical="top", horizontal="left")

        # Appliquer un style aux en-têtes (gras, centré horizontalement)
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center")

        print(f"[INFO] Formatage appliqué à la feuille '{sheet_name}'.")

    # Sauvegarde finale
    wb.save(file_path)
    print(f"\n[SUCCESS] Mise en forme terminée ! '{file_path}' est prêt à l'emploi.")

if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("[ERREUR] Aucun fichier Excel spécifié.")
        sys.exit(1)

    excel_file = sys.argv[1]
    format_excel(excel_file)
